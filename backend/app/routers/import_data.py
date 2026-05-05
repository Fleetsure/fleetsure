"""
Excel / CSV smart import router.
POST /import/preview  — upload file, get parsed preview + column mapping
POST /import/confirm  — submit mapped data, insert records into DB
"""

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.orm import Session
from sqlalchemy.exc import IntegrityError
from typing import List, Dict, Any, Optional, Tuple
from pydantic import BaseModel
import io
import openpyxl
from datetime import datetime, date
from decimal import Decimal, InvalidOperation

from app.database import get_db
from app.models.vehicle import Vehicle, VehicleType, VehicleStatus
from app.models.driver import Driver, DriverStatus, LicenseClass
from app.models.trip import Trip, TripStatus
from app.models.fuel_log import FuelLog
from app.models.user import User
from app.services.auth_service import get_current_user

router = APIRouter(prefix="/import", tags=["Import"])


# ── Schemas ────────────────────────────────────────────────────────────────────

class SheetPreview(BaseModel):
    sheet_name: str
    entity_type: str          # vehicles | drivers | trips | fuel | unknown
    headers: List[str]
    column_map: Dict[str, str]   # excel_column → our_field (or "" if unmapped)
    rows: List[Dict[str, Any]]   # ALL rows (frontend shows first 5, sends all on confirm)
    total_rows: int

class ImportConfirmSheet(BaseModel):
    sheet_name: str
    entity_type: str
    column_map: Dict[str, str]   # excel_column → our_field
    rows: List[Dict[str, Any]]   # ALL rows (raw values as strings)

class ImportConfirmRequest(BaseModel):
    sheets: List[ImportConfirmSheet]

class ImportResult(BaseModel):
    entity_type: str
    sheet_name: str
    inserted: int
    skipped: int
    errors: List[str]


# ── Keyword maps for fuzzy header matching ─────────────────────────────────────

VEHICLE_FIELDS: Dict[str, List[str]] = {
    "registration_number": ["reg", "registration", "vehicle no", "reg no", "plate", "truck no",
                            "number plate", "vehicle number", "veh no", "vehicle reg", "truck reg"],
    "make":         ["make", "manufacturer", "brand", "company"],
    "model":        ["model", "model name", "vehicle model"],
    "year":         ["year", "manufacture year", "mfg year", "yr", "manufacturing year"],
    "vehicle_type": ["type", "vehicle type", "veh type", "category"],
    "fuel_type":    ["fuel", "fuel type"],
    "chassis_number": ["chassis", "chassis no", "chassis number"],
    "engine_number":  ["engine", "engine no", "engine number"],
}

DRIVER_FIELDS: Dict[str, List[str]] = {
    "name":            ["name", "driver name", "driver", "full name"],
    "phone":           ["phone", "mobile", "contact", "number", "mob", "cell", "phone no", "mobile no"],
    "alternate_phone": ["alternate", "alt phone", "alt mobile", "other number", "alternate phone"],
    "license_number":  ["license", "dl no", "driving license", "dl number", "licence", "dl", "license no"],
    "license_expiry":  ["license expiry", "dl expiry", "expiry", "license exp", "dl exp", "validity", "license validity"],
    "address":         ["address", "addr", "location", "residence"],
    "blood_group":     ["blood", "blood group", "bg"],
    "dob":             ["dob", "date of birth", "birth date", "birthdate"],
}

TRIP_FIELDS: Dict[str, List[str]] = {
    "registration_number": ["vehicle", "truck", "reg no", "registration", "vehicle no", "truck no",
                            "veh", "vehicle reg", "truck reg"],
    "driver_name":      ["driver", "driver name"],
    "driver_phone":     ["driver phone", "driver mobile", "driver contact", "driver mob"],
    "origin":           ["from", "origin", "source", "from city", "pickup", "loading", "loading point"],
    "destination":      ["to", "destination", "to city", "delivery", "unloading", "unloading point"],
    "start_date":       ["start date", "start", "date", "trip date", "departure", "from date", "departure date"],
    "end_date":         ["end date", "end", "arrival", "to date", "completion", "arrival date"],
    "freight_amount":   ["freight", "amount", "revenue", "income", "freight amount", "billing",
                         "charge", "rate", "trip amount", "total amount"],
    "material":         ["material", "goods", "cargo", "commodity", "item", "product"],
    "weight_tonnes":    ["weight", "tonnes", "ton", "tons", "load", "weight tonnes"],
    "doc_number":       ["lr", "lr no", "lr number", "doc no", "consignment", "bill no", "lorry receipt"],
    "notes":            ["notes", "remarks", "comment", "comments"],
}

FUEL_FIELDS: Dict[str, List[str]] = {
    "registration_number": ["vehicle", "truck", "reg no", "registration", "vehicle no", "truck no"],
    "date":           ["date", "fill date", "fuel date", "filling date"],
    "litres":         ["litres", "liters", "qty", "quantity", "fuel qty", "volume", "ltr", "litre"],
    "amount":         ["amount", "total", "cost", "price", "total cost", "total amount"],
    "odometer_km":    ["odometer", "km", "reading", "odo", "odometer km", "km reading", "mileage", "odometer reading"],
    "fuel_station":   ["station", "pump", "fuel station", "petrol pump", "bunk", "filling station"],
    "notes":          ["notes", "remarks", "comment"],
}

ENTITY_SIGNALS: Dict[str, List[str]] = {
    "vehicles": ["registration", "make", "model", "vehicle type", "chassis"],
    "drivers":  ["license", "dl", "licence", "blood group", "driving"],
    "trips":    ["origin", "destination", "freight", "loading", "unloading", "lr"],
    "fuel":     ["litres", "liters", "fuel", "odometer", "pump", "bunk"],
}


# ── Utility functions ──────────────────────────────────────────────────────────

def normalize(s: str) -> str:
    return str(s).lower().strip().replace("_", " ").replace("-", " ")


def match_header(header: str, field_map: Dict[str, List[str]]) -> str:
    h = normalize(header)
    for field, keywords in field_map.items():
        for kw in keywords:
            if kw == h or kw in h or h in kw:
                return field
    return ""


def detect_entity(headers: List[str]) -> str:
    header_text = " ".join(normalize(h) for h in headers if h)
    scores = {entity: sum(1 for kw in kws if kw in header_text)
              for entity, kws in ENTITY_SIGNALS.items()}
    best = max(scores, key=scores.get)
    return best if scores[best] > 0 else "unknown"


def build_column_map(headers: List[str], entity_type: str) -> Dict[str, str]:
    field_map = {
        "vehicles": VEHICLE_FIELDS,
        "drivers":  DRIVER_FIELDS,
        "trips":    TRIP_FIELDS,
        "fuel":     FUEL_FIELDS,
    }.get(entity_type, {})
    return {h: match_header(h, field_map) for h in headers if h}


def parse_date(val) -> Optional[date]:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    if not s or s.lower() in ("none", "null", "nan", "-"):
        return None
    for fmt in ["%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y",
                "%d-%b-%Y", "%d %b %Y", "%d-%b-%y", "%d/%m/%y"]:
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def parse_num(val) -> Optional[Decimal]:
    if val is None:
        return None
    try:
        cleaned = str(val).replace(",", "").replace("₹", "").replace("Rs", "").strip()
        if not cleaned or cleaned.lower() in ("none", "null", "nan", "-"):
            return None
        return Decimal(cleaned)
    except InvalidOperation:
        return None


def parse_int(val) -> Optional[int]:
    n = parse_num(val)
    return int(n) if n is not None else None


def safe_str(val) -> Optional[str]:
    if val is None:
        return None
    s = str(val).strip()
    return s if s and s.lower() not in ("none", "null", "nan") else None


def parse_sheet_data(ws) -> Tuple[List[str], List[Dict]]:
    """Extract headers + rows from an openpyxl worksheet."""
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        return [], []

    # Find first non-empty row → treat as header
    header_idx = 0
    for i, row in enumerate(all_rows):
        if any(c is not None for c in row):
            header_idx = i
            break

    raw_headers = [str(c).strip() if c is not None else "" for c in all_rows[header_idx]]
    # Trim trailing empty columns
    while raw_headers and not raw_headers[-1]:
        raw_headers.pop()

    headers = raw_headers
    data_rows = []
    for row in all_rows[header_idx + 1:]:
        if not any(c is not None for c in row):
            continue  # skip blank rows
        row_dict = {}
        for i, h in enumerate(headers):
            row_dict[h] = row[i] if i < len(row) else None
        data_rows.append(row_dict)

    return headers, data_rows


# ── Preview endpoint ───────────────────────────────────────────────────────────

@router.post("/preview", response_model=List[SheetPreview])
async def preview_import(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
):
    content = await file.read()
    filename = (file.filename or "").lower()

    if not (filename.endswith(".xlsx") or filename.endswith(".xls") or filename.endswith(".csv")):
        raise HTTPException(400, "Only .xlsx, .xls, and .csv files are supported")

    previews: List[SheetPreview] = []

    # ── CSV ──────────────────────────────────────────────────────────────────
    if filename.endswith(".csv"):
        import csv as csv_mod
        text = content.decode("utf-8-sig", errors="replace")
        reader = csv_mod.DictReader(io.StringIO(text))
        headers = reader.fieldnames or []
        rows = [dict(r) for r in reader]
        entity_type = detect_entity(list(headers))
        col_map = build_column_map(list(headers), entity_type)
        all_rows = [
            {h: str(r.get(h, "") if r.get(h) is not None else "") for h in headers}
            for r in rows
        ]
        previews.append(SheetPreview(
            sheet_name="Sheet1",
            entity_type=entity_type,
            headers=list(headers),
            column_map=col_map,
            rows=all_rows,
            total_rows=len(rows),
        ))
        return previews

    # ── Excel ────────────────────────────────────────────────────────────────
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(400, f"Could not read Excel file: {e}")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers, rows = parse_sheet_data(ws)
        if not headers or not rows:
            continue
        entity_type = detect_entity(headers)
        col_map = build_column_map(headers, entity_type)
        all_rows = [
            {h: str(r.get(h, "") if r.get(h) is not None else "") for h in headers}
            for r in rows
        ]
        previews.append(SheetPreview(
            sheet_name=sheet_name,
            entity_type=entity_type,
            headers=headers,
            column_map=col_map,
            rows=all_rows,
            total_rows=len(rows),
        ))

    if not previews:
        raise HTTPException(400, "No usable data found in the file")

    return previews


# ── Confirm / insert endpoint ──────────────────────────────────────────────────

@router.post("/confirm", response_model=List[ImportResult])
def confirm_import(
    payload: ImportConfirmRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    results: List[ImportResult] = []

    for sheet in payload.sheets:
        if sheet.entity_type == "vehicles":
            r = _import_vehicles(sheet, db, current_user)
        elif sheet.entity_type == "drivers":
            r = _import_drivers(sheet, db, current_user)
        elif sheet.entity_type == "trips":
            r = _import_trips(sheet, db, current_user)
        elif sheet.entity_type == "fuel":
            r = _import_fuel(sheet, db, current_user)
        else:
            r = ImportResult(entity_type=sheet.entity_type, sheet_name=sheet.sheet_name,
                             inserted=0, skipped=0, errors=["Unknown entity type — skipped"])
        results.append(r)

    return results


# ── Per-entity importers ───────────────────────────────────────────────────────

def _mapped(row: Dict, col_map: Dict[str, str], field: str) -> Any:
    """Get raw value from row for our field, using reverse col_map lookup."""
    for col, mapped in col_map.items():
        if mapped == field and col in row:
            return row[col]
    return None


def _import_vehicles(sheet: ImportConfirmSheet, db: Session, user: User) -> ImportResult:
    inserted = skipped = 0
    errors: List[str] = []
    m = sheet.column_map

    for i, row in enumerate(sheet.rows, start=2):
        reg = safe_str(_mapped(row, m, "registration_number"))
        if not reg:
            skipped += 1
            continue

        reg = reg.upper().replace(" ", "")
        existing = db.query(Vehicle).filter(Vehicle.registration_number == reg).first()
        if existing:
            skipped += 1
            continue

        make = safe_str(_mapped(row, m, "make")) or "Unknown"
        model = safe_str(_mapped(row, m, "model")) or "Unknown"

        vtype_raw = safe_str(_mapped(row, m, "vehicle_type"))
        vtype = VehicleType.TRUCK
        if vtype_raw:
            vt = vtype_raw.lower()
            if "trailer" in vt:
                vtype = VehicleType.TRAILER
            elif "tanker" in vt:
                vtype = VehicleType.TANKER
            elif "container" in vt:
                vtype = VehicleType.CONTAINER
            elif "mini" in vt:
                vtype = VehicleType.MINI_TRUCK

        try:
            v = Vehicle(
                registration_number=reg,
                make=make,
                model=model,
                year=parse_int(_mapped(row, m, "year")),
                vehicle_type=vtype,
                fuel_type=safe_str(_mapped(row, m, "fuel_type")),
                chassis_number=safe_str(_mapped(row, m, "chassis_number")),
                engine_number=safe_str(_mapped(row, m, "engine_number")),
                owner_id=user.id,
            )
            db.add(v)
            db.flush()
            inserted += 1
        except Exception as e:
            errors.append(f"Row {i} ({reg}): {e}")
            db.rollback()

    db.commit()
    return ImportResult(entity_type="vehicles", sheet_name=sheet.sheet_name,
                        inserted=inserted, skipped=skipped, errors=errors)


def _import_drivers(sheet: ImportConfirmSheet, db: Session, user: User) -> ImportResult:
    inserted = skipped = 0
    errors: List[str] = []
    m = sheet.column_map

    for i, row in enumerate(sheet.rows, start=2):
        name = safe_str(_mapped(row, m, "name"))
        phone = safe_str(_mapped(row, m, "phone"))
        if not name or not phone:
            skipped += 1
            continue

        # Normalize phone: keep digits only, ensure 10 digits
        phone_digits = "".join(filter(str.isdigit, phone))
        if len(phone_digits) >= 10:
            phone = phone_digits[-10:]
        else:
            phone = phone_digits

        existing = db.query(Driver).filter(Driver.phone == phone).first()
        if existing:
            skipped += 1
            continue

        dl_num = safe_str(_mapped(row, m, "license_number"))
        if dl_num:
            existing_dl = db.query(Driver).filter(Driver.license_number == dl_num).first()
            if existing_dl:
                skipped += 1
                continue

        try:
            d = Driver(
                name=name,
                phone=phone,
                alternate_phone=safe_str(_mapped(row, m, "alternate_phone")),
                address=safe_str(_mapped(row, m, "address")),
                license_number=dl_num,
                license_expiry=parse_date(_mapped(row, m, "license_expiry")),
                dob=parse_date(_mapped(row, m, "dob")),
                blood_group=safe_str(_mapped(row, m, "blood_group")),
                owner_id=user.id,
            )
            db.add(d)
            db.flush()
            inserted += 1
        except Exception as e:
            errors.append(f"Row {i} ({name}): {e}")
            db.rollback()

    db.commit()
    return ImportResult(entity_type="drivers", sheet_name=sheet.sheet_name,
                        inserted=inserted, skipped=skipped, errors=errors)


def _import_trips(sheet: ImportConfirmSheet, db: Session, user: User) -> ImportResult:
    inserted = skipped = 0
    errors: List[str] = []
    m = sheet.column_map

    for i, row in enumerate(sheet.rows, start=2):
        reg = safe_str(_mapped(row, m, "registration_number"))
        origin = safe_str(_mapped(row, m, "origin"))
        destination = safe_str(_mapped(row, m, "destination"))
        start_date = parse_date(_mapped(row, m, "start_date"))

        if not reg or not origin or not destination or not start_date:
            skipped += 1
            continue

        reg = reg.upper().replace(" ", "")
        vehicle = db.query(Vehicle).filter(
            Vehicle.registration_number == reg,
            Vehicle.owner_id == user.id
        ).first()
        if not vehicle:
            errors.append(f"Row {i}: Vehicle '{reg}' not found — add vehicle first")
            skipped += 1
            continue

        freight = parse_num(_mapped(row, m, "freight_amount")) or Decimal("0")

        try:
            t = Trip(
                vehicle_id=vehicle.id,
                driver_name=safe_str(_mapped(row, m, "driver_name")) or "Unknown",
                driver_phone=safe_str(_mapped(row, m, "driver_phone")),
                origin=origin,
                destination=destination,
                start_date=start_date,
                end_date=parse_date(_mapped(row, m, "end_date")),
                freight_amount=freight,
                material=safe_str(_mapped(row, m, "material")),
                weight_tonnes=parse_num(_mapped(row, m, "weight_tonnes")),
                doc_number=safe_str(_mapped(row, m, "doc_number")),
                notes=safe_str(_mapped(row, m, "notes")),
                status=TripStatus.COMPLETED,
                owner_id=user.id,
            )
            db.add(t)
            db.flush()
            inserted += 1
        except Exception as e:
            errors.append(f"Row {i}: {e}")
            db.rollback()

    db.commit()
    return ImportResult(entity_type="trips", sheet_name=sheet.sheet_name,
                        inserted=inserted, skipped=skipped, errors=errors)


def _import_fuel(sheet: ImportConfirmSheet, db: Session, user: User) -> ImportResult:
    inserted = skipped = 0
    errors: List[str] = []
    m = sheet.column_map

    for i, row in enumerate(sheet.rows, start=2):
        reg = safe_str(_mapped(row, m, "registration_number"))
        fuel_date = parse_date(_mapped(row, m, "date"))
        litres = parse_num(_mapped(row, m, "litres"))
        amount = parse_num(_mapped(row, m, "amount"))

        if not reg or not fuel_date or not litres or not amount:
            skipped += 1
            continue

        reg = reg.upper().replace(" ", "")
        vehicle = db.query(Vehicle).filter(
            Vehicle.registration_number == reg,
            Vehicle.owner_id == user.id
        ).first()
        if not vehicle:
            errors.append(f"Row {i}: Vehicle '{reg}' not found — add vehicle first")
            skipped += 1
            continue

        try:
            f = FuelLog(
                vehicle_id=vehicle.id,
                date=fuel_date,
                litres=litres,
                amount=amount,
                odometer_km=parse_num(_mapped(row, m, "odometer_km")),
                fuel_station=safe_str(_mapped(row, m, "fuel_station")),
                notes=safe_str(_mapped(row, m, "notes")),
                owner_id=user.id,
            )
            db.add(f)
            db.flush()
            inserted += 1
        except Exception as e:
            errors.append(f"Row {i}: {e}")
            db.rollback()

    db.commit()
    return ImportResult(entity_type="fuel", sheet_name=sheet.sheet_name,
                        inserted=inserted, skipped=skipped, errors=errors)
