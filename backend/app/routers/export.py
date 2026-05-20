"""
Export router — download fleet data as XLSX or CSV (ZIP).
Supports: vehicles, drivers, trips, fuel, tolls, tyres, misc, profit_loss
"""
import csv
import io
import zipfile
from datetime import datetime, date
from decimal import Decimal

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse

from app.db import supabase
from app.services.auth_service import get_current_user

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False

router = APIRouter(prefix="/export", tags=["export"])


def _safe(val):
    if val is None:
        return ""
    if isinstance(val, Decimal):
        return float(val)
    if isinstance(val, (datetime, date)):
        return str(val)
    return val


def _build_datasets(types: list[str], owner_id: str) -> dict:
    result = {}

    if "vehicles" in types:
        rows = supabase.table("vehicles").select("*").eq("owner_id", owner_id).execute().data
        result["Vehicles"] = (
            ["Registration No", "Make", "Model", "Year", "Type", "Fuel Type", "Status",
             "Insurance Expiry", "Fitness Expiry", "PUC Expiry", "Permit Expiry"],
            [[_safe(v.get("registration_number")), _safe(v.get("make")), _safe(v.get("model")),
              _safe(v.get("year")), _safe(v.get("vehicle_type")), _safe(v.get("fuel_type")),
              _safe(v.get("status")), _safe(v.get("insurance_expiry")), _safe(v.get("fitness_expiry")),
              _safe(v.get("puc_expiry")), _safe(v.get("permit_expiry"))] for v in rows]
        )

    if "drivers" in types:
        rows = supabase.table("drivers").select("*").eq("owner_id", owner_id).execute().data
        result["Drivers"] = (
            ["Name", "Phone", "Alt Phone", "License No", "License Class", "License Expiry",
             "Transport Valid Till", "Blood Group", "Status"],
            [[_safe(d.get("name")), _safe(d.get("phone")), _safe(d.get("alternate_phone")),
              _safe(d.get("license_number")), _safe(d.get("license_class")), _safe(d.get("license_expiry")),
              _safe(d.get("transport_validity")), _safe(d.get("blood_group")), _safe(d.get("status"))] for d in rows]
        )

    if "trips" in types:
        rows = supabase.table("trips").select("*").eq("owner_id", owner_id).order("start_date", desc=True).execute().data
        result["Trips"] = (
            ["Date", "From", "To", "Driver", "Driver Phone", "Vehicle",
             "Distance (km)", "Freight (₹)", "Status", "Notes"],
            [[_safe(t.get("start_date")), _safe(t.get("origin")), _safe(t.get("destination")),
              _safe(t.get("driver_name")), _safe(t.get("driver_phone")), _safe(t.get("vehicle_id")),
              _safe(t.get("distance_km")), _safe(t.get("freight_amount")),
              _safe(t.get("status")), _safe(t.get("notes"))] for t in rows]
        )

    if "fuel" in types:
        rows = supabase.table("fuel_logs").select("*").eq("owner_id", owner_id).order("date", desc=True).execute().data
        vehicles = supabase.table("vehicles").select("id,registration_number").eq("owner_id", owner_id).execute().data
        veh_map = {v["id"]: v["registration_number"] for v in vehicles}
        result["Fuel Log"] = (
            ["Date", "Vehicle", "Litres", "Amount (₹)", "Odometer (km)", "Fuel Station", "Notes"],
            [[_safe(f.get("date")), veh_map.get(f.get("vehicle_id"), ""), _safe(f.get("litres")),
              _safe(f.get("amount")), _safe(f.get("odometer_km")), _safe(f.get("fuel_station")), _safe(f.get("notes"))] for f in rows]
        )

    if "tolls" in types:
        rows = supabase.table("toll_logs").select("*").eq("owner_id", owner_id).order("date", desc=True).execute().data
        vehicles = supabase.table("vehicles").select("id,registration_number").eq("owner_id", owner_id).execute().data
        veh_map = {v["id"]: v["registration_number"] for v in vehicles}
        result["Tolls"] = (
            ["Date", "Vehicle", "Toll Plaza", "Route", "Payment Mode", "Amount (₹)", "Notes"],
            [[_safe(t.get("date")), veh_map.get(t.get("vehicle_id"), ""), _safe(t.get("toll_plaza")),
              _safe(t.get("route")), _safe(t.get("payment_mode")), _safe(t.get("amount")), _safe(t.get("notes"))] for t in rows]
        )

    if "tyres" in types:
        rows = supabase.table("tyre_logs").select("*").eq("owner_id", owner_id).order("date", desc=True).execute().data
        vehicles = supabase.table("vehicles").select("id,registration_number").eq("owner_id", owner_id).execute().data
        veh_map = {v["id"]: v["registration_number"] for v in vehicles}
        result["Tyres"] = (
            ["Date", "Vehicle", "Type", "Brand", "Count", "Position", "Odometer (km)", "Amount (₹)", "Notes"],
            [[_safe(t.get("date")), veh_map.get(t.get("vehicle_id"), ""), _safe(t.get("tyre_type")),
              _safe(t.get("tyre_brand")), _safe(t.get("tyre_count")), _safe(t.get("tyre_position")),
              _safe(t.get("odometer_km")), _safe(t.get("amount")), _safe(t.get("notes"))] for t in rows]
        )

    if "misc" in types:
        rows = supabase.table("misc_expenses").select("*").eq("owner_id", owner_id).order("date", desc=True).execute().data
        vehicles = supabase.table("vehicles").select("id,registration_number").eq("owner_id", owner_id).execute().data
        veh_map = {v["id"]: v["registration_number"] for v in vehicles}
        result["Misc Expenses"] = (
            ["Date", "Vehicle", "Category", "Description", "Amount (₹)", "Notes"],
            [[_safe(m.get("date")), veh_map.get(m.get("vehicle_id"), "") if m.get("vehicle_id") else "",
              _safe(m.get("category")), _safe(m.get("description")), _safe(m.get("amount")), _safe(m.get("notes"))] for m in rows]
        )

    if "profit_loss" in types:
        trips = supabase.table("trips").select("*").eq("owner_id", owner_id).order("start_date", desc=True).execute().data
        vehicles = supabase.table("vehicles").select("id,registration_number").eq("owner_id", owner_id).execute().data
        veh_map = {v["id"]: v["registration_number"] for v in vehicles}
        pl_rows = []
        for t in trips:
            freight = float(t.get("freight_amount") or 0)
            exps = supabase.table("expenses").select("expense_type,amount").eq("trip_id", t["id"]).execute().data
            total_exp = sum(float(e.get("amount") or 0) for e in exps)
            profit = freight - total_exp
            margin = round(profit / freight * 100, 1) if freight else 0
            fuel_exp = sum(float(e.get("amount") or 0) for e in exps if e.get("expense_type") == "fuel")
            toll_exp = sum(float(e.get("amount") or 0) for e in exps if e.get("expense_type") == "toll")
            maint_exp = sum(float(e.get("amount") or 0) for e in exps if e.get("expense_type") == "maintenance")
            drv_exp = sum(float(e.get("amount") or 0) for e in exps if e.get("expense_type") == "driver_payment")
            pl_rows.append([
                _safe(t.get("start_date")), _safe(t.get("origin")), _safe(t.get("destination")),
                veh_map.get(t.get("vehicle_id"), ""), _safe(t.get("driver_name")),
                freight, total_exp, profit, f"{margin}%",
                fuel_exp, toll_exp, maint_exp, drv_exp, _safe(t.get("status"))
            ])
        result["Profit & Loss"] = (
            ["Date", "From", "To", "Vehicle", "Driver", "Freight (₹)", "Total Expenses (₹)",
             "Profit (₹)", "Margin", "Fuel (₹)", "Toll (₹)", "Maintenance (₹)",
             "Driver Pay (₹)", "Status"],
            pl_rows
        )

    return result


def _style_header_row(ws, header_row: int, col_count: int):
    header_fill = PatternFill("solid", fgColor="1E2D8E")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    for col in range(1, col_count + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center")


def _auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)


def export_xlsx(datasets: dict, org_name: str) -> StreamingResponse:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws_sum = wb.create_sheet("Summary")
    ws_sum["A1"] = f"{org_name} — FleetSure Data Export"
    ws_sum["A1"].font = Font(bold=True, size=14, color="1E2D8E")
    ws_sum["A2"] = f"Generated: {datetime.now().strftime('%d %b %Y, %I:%M %p')}"
    ws_sum["A2"].font = Font(size=10, color="888888")
    ws_sum["A4"] = "Sheets in this workbook:"
    ws_sum["A4"].font = Font(bold=True)
    for i, sheet_name in enumerate(datasets.keys(), start=5):
        ws_sum[f"A{i}"] = f"  • {sheet_name}"
    _auto_width(ws_sum)

    for sheet_name, (headers, rows) in datasets.items():
        ws = wb.create_sheet(sheet_name)
        ws.append([sheet_name])
        ws["A1"].font = Font(bold=True, size=12, color="1E2D8E")
        ws.append([])
        ws.append(headers)
        _style_header_row(ws, 3, len(headers))
        ws.row_dimensions[3].height = 22
        alt_fill = PatternFill("solid", fgColor="F8F9FF")
        for r_idx, row in enumerate(rows):
            ws.append(row)
            if r_idx % 2 == 0:
                for col in range(1, len(headers) + 1):
                    ws.cell(row=4 + r_idx, column=col).fill = alt_fill
        ws.freeze_panes = "A4"
        _auto_width(ws)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"fleetsure_export_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"}
    )


def export_csv_zip(datasets: dict) -> StreamingResponse:
    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for sheet_name, (headers, rows) in datasets.items():
            csv_buf = io.StringIO()
            writer = csv.writer(csv_buf)
            writer.writerow(headers)
            writer.writerows(rows)
            safe_name = sheet_name.replace(" ", "_").replace("&", "and").lower()
            zf.writestr(f"{safe_name}.csv", csv_buf.getvalue())
    zip_buf.seek(0)

    fname = f"fleetsure_export_{datetime.now().strftime('%Y%m%d_%H%M')}.zip"
    return StreamingResponse(
        zip_buf,
        media_type="application/zip",
        headers={"Content-Disposition": f"attachment; filename={fname}"}
    )


@router.get("/")
def export_data(
    format: str = Query("xlsx"),
    types: str = Query("vehicles,drivers,trips,fuel,tolls,tyres,misc,profit_loss"),
    org_name: str = Query("My Fleet"),
    current_user: dict = Depends(get_current_user),
):
    selected = [t.strip().lower() for t in types.split(",") if t.strip()]
    datasets = _build_datasets(selected, current_user["id"])

    if not datasets:
        return {"error": "No data found"}

    if format == "csv":
        return export_csv_zip(datasets)

    if not XLSX_AVAILABLE:
        return {"error": "openpyxl not installed on server"}

    return export_xlsx(datasets, org_name)
